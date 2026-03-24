from flask import Flask, render_template, request, jsonify
import os
import json
from datetime import datetime

try:
    from openpyxl import Workbook, load_workbook
    EXCEL_EXPORT_AVAILABLE = True
except ImportError:
    Workbook = None
    load_workbook = None
    EXCEL_EXPORT_AVAILABLE = False

try:
    import psycopg
    from psycopg.rows import dict_row
    from psycopg.types.json import Json
    POSTGRES_AVAILABLE = True
except ImportError:
    psycopg = None
    dict_row = None
    Json = None
    POSTGRES_AVAILABLE = False

APP_DIR = os.path.dirname(__file__)
SUB_FILE = os.path.join(APP_DIR, 'submissions.json')
DESKTOP_DIR = os.path.join(os.path.expanduser('~'), 'Desktop')
EXCEL_FILE = os.path.join(DESKTOP_DIR, 'feedback_submissions.xlsx')
EXCEL_HEADERS = ['name', 'email', 'course', 'rating', 'comments', 'consent', 'receivedAt']
MAX_SUBMISSIONS_PER_PERSON = 2
DATABASE_URL = os.environ.get('DATABASE_URL', '').strip()
VERCEL_ENV = os.environ.get('VERCEL', '').strip()


def normalize_identity(value):
    return (value or '').strip().lower()


def use_postgres():
    return bool(DATABASE_URL and POSTGRES_AVAILABLE)


def should_use_excel_export():
    return EXCEL_EXPORT_AVAILABLE and not VERCEL_ENV


def get_db_connection():
    if not use_postgres():
        return None
    return psycopg.connect(DATABASE_URL, row_factory=dict_row)


def ensure_database():
    if not use_postgres():
        return

    with get_db_connection() as conn:
        with conn.cursor() as cur:
            cur.execute(
                '''
                CREATE TABLE IF NOT EXISTS feedback_submissions (
                    id BIGSERIAL PRIMARY KEY,
                    name TEXT NOT NULL,
                    email TEXT,
                    course TEXT NOT NULL,
                    teachers JSONB NOT NULL DEFAULT '[]'::jsonb,
                    teacher_comments JSONB NOT NULL DEFAULT '{}'::jsonb,
                    rating TEXT NOT NULL,
                    comments TEXT,
                    consent BOOLEAN NOT NULL DEFAULT FALSE,
                    received_at TIMESTAMPTZ NOT NULL
                )
                '''
            )
        conn.commit()


def ensure_submissions_file():
    if not os.path.exists(SUB_FILE):
        with open(SUB_FILE, 'w', encoding='utf-8') as f:
            json.dump([], f)


def load_file_submissions():
    ensure_submissions_file()
    try:
        with open(SUB_FILE, 'r', encoding='utf-8') as f:
            return json.load(f)
    except Exception:
        return []


def save_file_submission(entry):
    arr = load_file_submissions()
    arr.append(entry)
    with open(SUB_FILE, 'w', encoding='utf-8') as f:
        json.dump(arr, f, indent=2, ensure_ascii=False)


def get_postgres_submissions():
    ensure_database()
    with get_db_connection() as conn:
        with conn.cursor() as cur:
            cur.execute(
                '''
                SELECT
                    name,
                    email,
                    course,
                    teachers,
                    teacher_comments AS "teacherComments",
                    rating,
                    comments,
                    consent,
                    to_char(received_at AT TIME ZONE 'UTC', 'YYYY-MM-DD"T"HH24:MI:SS.MS"Z"') AS "receivedAt"
                FROM feedback_submissions
                ORDER BY received_at DESC
                '''
            )
            return cur.fetchall()


def count_postgres_submissions(name, email):
    ensure_database()
    with get_db_connection() as conn:
        with conn.cursor() as cur:
            cur.execute(
                '''
                SELECT COUNT(*) AS total
                FROM feedback_submissions
                WHERE lower(trim(name)) = %s
                  AND lower(trim(COALESCE(email, ''))) = %s
                ''',
                (normalize_identity(name), normalize_identity(email))
            )
            row = cur.fetchone()
            return row['total'] if row else 0


def save_postgres_submission(entry):
    ensure_database()
    with get_db_connection() as conn:
        with conn.cursor() as cur:
            cur.execute(
                '''
                INSERT INTO feedback_submissions (
                    name,
                    email,
                    course,
                    teachers,
                    teacher_comments,
                    rating,
                    comments,
                    consent,
                    received_at
                )
                VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s)
                ''',
                (
                    entry.get('name', ''),
                    entry.get('email', ''),
                    entry.get('course', ''),
                    Json(entry.get('teachers', [])),
                    Json(entry.get('teacherComments', {})),
                    entry.get('rating', ''),
                    entry.get('comments', ''),
                    bool(entry.get('consent')),
                    entry.get('receivedAt')
                )
            )
        conn.commit()


def count_matching_submissions(entries, name, email):
    normalized_name = normalize_identity(name)
    normalized_email = normalize_identity(email)
    return sum(
        1
        for entry in entries
        if normalize_identity(entry.get('name')) == normalized_name
        and normalize_identity(entry.get('email')) == normalized_email
    )


def get_submission_count(name, email):
    if use_postgres():
        return count_postgres_submissions(name, email)
    return count_matching_submissions(load_file_submissions(), name, email)


def get_all_submissions():
    if use_postgres():
        return get_postgres_submissions()
    return load_file_submissions()


def save_submission(entry):
    if use_postgres():
        save_postgres_submission(entry)
        return
    save_file_submission(entry)


def ensure_excel_file():
    if not should_use_excel_export():
        return

    os.makedirs(DESKTOP_DIR, exist_ok=True)
    if os.path.exists(EXCEL_FILE):
        return

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = 'Feedback'
    sheet.append(EXCEL_HEADERS)
    for entry in get_all_submissions():
        sheet.append([entry.get(header, '') for header in EXCEL_HEADERS])
    workbook.save(EXCEL_FILE)


def append_submission_to_excel(entry):
    if not should_use_excel_export():
        return

    ensure_excel_file()
    workbook = load_workbook(EXCEL_FILE)
    sheet = workbook.active
    sheet.append([entry.get(header, '') for header in EXCEL_HEADERS])
    workbook.save(EXCEL_FILE)


app = Flask(__name__, static_folder='static', template_folder='templates')


@app.route('/')
def index():
    return render_template('index.html')


@app.route('/api/health', methods=['GET'])
def health():
    return jsonify({
        'ok': True,
        'status': 'healthy',
        'storage': 'postgres' if use_postgres() else 'json'
    })


@app.route('/submit', methods=['POST'])
def submit():
    data = request.get_json() or {}
    if not data.get('name') or not data.get('course') or not data.get('rating'):
        return jsonify({'error': 'name, course and rating are required'}), 400

    if get_submission_count(data.get('name'), data.get('email')) >= MAX_SUBMISSIONS_PER_PERSON:
        return jsonify({'error': 'This name and email can only submit feedback twice.'}), 400

    entry = dict(data)
    entry['teachers'] = entry.get('teachers') or []
    entry['teacherComments'] = entry.get('teacherComments') or {}
    entry['consent'] = bool(entry.get('consent'))
    entry['receivedAt'] = datetime.utcnow().isoformat() + 'Z'

    save_submission(entry)
    try:
        append_submission_to_excel(entry)
    except Exception as exc:
        print(f'Excel export skipped: {exc}')

    return jsonify({'ok': True, 'storage': 'postgres' if use_postgres() else 'json'})


@app.route('/submissions', methods=['GET'])
def submissions():
    return jsonify(get_all_submissions())


if __name__ == '__main__':
    if use_postgres():
        ensure_database()
    else:
        ensure_submissions_file()
    try:
        ensure_excel_file()
    except Exception as exc:
        print(f'Excel setup skipped: {exc}')
    app.run(host='127.0.0.1', port=5000, debug=True, use_reloader=False)
